from django.test import TestCase, Client
from django.contrib.auth import get_user_model
from django.urls import reverse
import openpyxl
import io

from core.models import Modalidade, ChaveamentoModalidade, PartidaChaveamento, Jogo
from exportador.services import gerar_planilha_jogos_xlsx

User = get_user_model()


class ExportadorExcelTests(TestCase):
    def setUp(self):
        self.client = Client()
        
        # Usuário Comissão
        self.comissao_user = User.objects.create_user(
            email='comissao@test.com',
            password='password123',
            role='COMISSAO',
            perfil_completo=True,
            nome_completo='Membro Comissão'
        )
        
        # Delegações de teste
        self.delegacao_a = User.objects.create_user(
            email='campus_a@test.com',
            password='password123',
            role='REPRESENTANTE',
            perfil_completo=True,
            nome_delegacao='Campus Diamantina'
        )
        self.delegacao_b = User.objects.create_user(
            email='campus_b@test.com',
            password='password123',
            role='REPRESENTANTE',
            perfil_completo=True,
            nome_delegacao='Campus Mucuri'
        )

        # Modalidade e Chaveamento de teste
        self.modalidade = Modalidade.objects.create(
            nome='Futsal',
            genero='M'
        )
        self.chaveamento = ChaveamentoModalidade.objects.create(
            modalidade=self.modalidade,
            fase_atual='fase_grupos'
        )
        
        # Partida com delegações
        self.partida_1 = PartidaChaveamento.objects.create(
            chaveamento=self.chaveamento,
            fase='GRUPO_LOCAL',
            rodada=1,
            time_a=self.delegacao_a,
            time_b=self.delegacao_b
        )
        
        # Partida de Mata-Mata sem delegações ainda (A Definir)
        self.partida_2 = PartidaChaveamento.objects.create(
            chaveamento=self.chaveamento,
            fase='FINAL_LOCAL',
            rodada=2,
            time_a=None,
            time_b=None
        )

    def test_acesso_negado_para_usuario_nao_autenticado(self):
        url = reverse('exportador:exportar_jogos_excel')
        response = self.client.get(url)
        self.assertNotEqual(response.status_code, 200)

    def test_acesso_negado_para_usuario_comum(self):
        self.client.force_login(self.delegacao_a)
        url = reverse('exportador:exportar_jogos_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)

    def test_exportacao_excel_sucesso_comissao(self):
        self.client.force_login(self.comissao_user)
        url = reverse('exportador:exportar_jogos_excel')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment; filename=', response['Content-Disposition'])
        
        # Valida abertura do Excel gerado na memória
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Resumo Geral", wb.sheetnames)
        self.assertIn("Todos os Jogos", wb.sheetnames)
        
        # Verifica se os dados das partidas e delegações estão na planilha
        ws_master = wb["Todos os Jogos"]
        val_cell_a = ws_master["F3"].value
        val_cell_b = ws_master["G3"].value
        self.assertEqual(val_cell_a, 'Campus Diamantina')
        self.assertEqual(val_cell_b, 'Campus Mucuri')
        
        # Verifica vaga a definir do mata-mata na linha 4
        val_cell_final = ws_master["F4"].value
        self.assertIn('A definir', str(val_cell_final))

    def test_exportacao_filtrada_por_modalidade(self):
        self.client.force_login(self.comissao_user)
        url = f"{reverse('exportador:exportar_jogos_excel')}?modalidade_id={self.modalidade.id}"
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Resumo Geral", wb.sheetnames)
