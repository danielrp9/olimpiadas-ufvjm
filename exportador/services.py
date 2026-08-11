import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import Modalidade, ChaveamentoModalidade, PartidaChaveamento, Jogo


def formatar_time_display(time_obj, partida, posicao='a'):
    """
    Retorna o nome formatado da delegação/time para a planilha.
    Se o time ainda não estiver definido no banco (mata-mata pendente),
    retorna uma descrição amigável da vaga no chaveamento.
    """
    if time_obj:
        return time_obj.nome_delegacao or time_obj.nome_completo or time_obj.email
    
    # Se o time não está definido no banco, geramos a descrição da vaga
    fase = partida.fase if partida else ''
    if fase == 'QUARTAS_LOCAL':
        return f"A definir (Quartas de Final - Time {posicao.upper()})"
    elif fase in ['SEMI_LOCAL', 'SEMI_GERAL']:
        return f"A definir (Semifinalista - Time {posicao.upper()})"
    elif fase in ['FINAL_LOCAL', 'FINAL_GERAL']:
        return f"A definir (Finalista - Time {posicao.upper()})"
    elif fase in ['DISPUTA_3_LOCAL', 'BRONZE']:
        return f"A definir (Disputa 3º Lugar - Time {posicao.upper()})"
    elif fase == 'EXTERNO_ELIMINATORIA':
        return f"A definir (Eliminatória Externa - Time {posicao.upper()})"
    return f"A definir (Time {posicao.upper()})"


def gerar_planilha_jogos_xlsx(modalidade_id=None):
    """
    Gera uma Pasta de Trabalho Excel (.xlsx) contendo:
    1. Aba de Resumo Geral (Dashboard)
    2. Aba com a Lista Unificada de Todos os Jogos
    3. Abas individuais por Modalidade Esportiva
    
    Retorna o conteúdo binário (bytes) do arquivo XLSX.
    """
    wb = openpyxl.Workbook()
    
    # Configura Aba 1 (Resumo)
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Geral"
    
    # --- ESTILOS VISUAIS ---
    header_fill_dark = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Slate 800
    header_fill_blue = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")  # Blue 700
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")        # Slate 50
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")        # Slate 200
    
    font_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="E2E8F0")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_regular = Font(name="Calibri", size=11, color="0F172A")
    font_small = Font(name="Calibri", size=10, italic=True, color="64748B")
    
    thin_border_side = Side(style='thin', color='CBD5E1')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(
        top=Side(style='thin', color='475569'),
        bottom=Side(style='double', color='1E293B'),
        left=thin_border_side,
        right=thin_border_side
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Filtragem de Modalidades
    modalidades_qs = Modalidade.objects.all().order_by('nome')
    if modalidade_id:
        modalidades_qs = modalidades_qs.filter(id=modalidade_id)
        
    # --- 1. CONSTRUÇÃO DA ABA RESUMO GERAL ---
    ws_resumo.views.sheetView[0].showGridLines = True
    ws_resumo.merge_cells("A1:G1")
    title_cell = ws_resumo["A1"]
    title_cell.value = "OLIMPÍADAS UNIVERSITÁRIAS - PAINEL GERAL DE JOGOS E CHAVEAMENTOS"
    title_cell.font = font_title
    title_cell.fill = header_fill_dark
    title_cell.alignment = align_center
    ws_resumo.row_dimensions[1].height = 36

    ws_resumo.merge_cells("A2:G2")
    sub_cell = ws_resumo["A2"]
    sub_cell.value = "Planilha de Rascunho para Planejamento Operacional, Datas, Horários e Quadras"
    sub_cell.font = font_subtitle
    sub_cell.fill = header_fill_dark
    sub_cell.alignment = align_center
    ws_resumo.row_dimensions[2].height = 22

    ws_resumo.append([])  # Linha em branco (3)

    headers_resumo = [
        "Modalidade", "Gênero", "Status Chaveamento", "Grupos Criados",
        "Total de Jogos", "Jogos Agendados", "Jogos Pendentes"
    ]
    ws_resumo.append(headers_resumo)
    ws_resumo.row_dimensions[4].height = 26
    
    for col_idx, text in enumerate(headers_resumo, start=1):
        cell = ws_resumo.cell(row=4, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill_blue
        cell.alignment = align_center
        cell.border = border_all

    total_jogos_geral = 0
    total_agendados_geral = 0
    total_pendentes_geral = 0
    total_grupos_geral = 0
    
    row_idx = 5
    modalidades_com_partidas = []

    for mod in modalidades_qs:
        ch = getattr(mod, 'chaveamento', None)
        status_ch = ch.get_fase_atual_display() if ch else "Não Gerado"
        num_grupos = ch.grupos.count() if ch else 0
        
        partidas_mod = list(PartidaChaveamento.objects.filter(chaveamento=ch).select_related('time_a', 'time_b', 'jogo', 'grupo')) if ch else []
        jogos_diretos = list(Jogo.objects.filter(modalidade=mod, partida_chaveamento__isnull=True).select_related('time_a', 'time_b'))
        
        all_matches_count = len(partidas_mod) + len(jogos_diretos)
        if all_matches_count > 0:
            modalidades_com_partidas.append((mod, ch, partidas_mod, jogos_diretos))
            
        agendados_count = 0
        for p in partidas_mod:
            if p.data_partida or (p.jogo and p.jogo.data_jogo):
                agendados_count += 1
        for j in jogos_diretos:
            if j.data_jogo:
                agendados_count += 1
                
        pendentes_count = all_matches_count - agendados_count
        
        total_jogos_geral += all_matches_count
        total_agendados_geral += agendados_count
        total_pendentes_geral += pendentes_count
        total_grupos_geral += num_grupos
        
        row_data = [
            mod.nome,
            mod.get_genero_display(),
            status_ch,
            num_grupos,
            all_matches_count,
            agendados_count,
            pendentes_count
        ]
        
        ws_resumo.append(row_data)
        ws_resumo.row_dimensions[row_idx].height = 22
        
        fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx in range(1, 8):
            cell = ws_resumo.cell(row=row_idx, column=c_idx)
            cell.font = font_regular
            cell.border = border_all
            if fill.fill_type:
                cell.fill = fill
            if c_idx in [2, 3, 4, 5, 6, 7]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        row_idx += 1

    # Linha de Totalização no Resumo
    total_row = [
        "TOTAL GERAL", "", "", total_grupos_geral,
        total_jogos_geral, total_agendados_geral, total_pendentes_geral
    ]
    ws_resumo.append(total_row)
    ws_resumo.row_dimensions[row_idx].height = 26
    for c_idx in range(1, 8):
        cell = ws_resumo.cell(row=row_idx, column=c_idx)
        cell.font = font_bold
        cell.fill = total_fill
        cell.border = border_total
        if c_idx in [4, 5, 6, 7]:
            cell.alignment = align_center

    # --- 2. CONSTRUÇÃO DA ABA LISTA MESTRE (TODOS OS JOGOS) ---
    ws_master = wb.create_sheet(title="Todos os Jogos")
    ws_master.views.sheetView[0].showGridLines = True
    
    ws_master.merge_cells("A1:L1")
    m_title = ws_master["A1"]
    m_title.value = "VISÃO GERAL UNIFICADA DE TODOS OS JOGOS"
    m_title.font = font_title
    m_title.fill = header_fill_dark
    m_title.alignment = align_center
    ws_master.row_dimensions[1].height = 36

    headers_master = [
        "# Jogo", "Modalidade", "Fase", "Grupo / Chave", "Rodada",
        "Time A (Delegação)", "Time B (Delegação)",
        "Data do Jogo", "Horário", "Local / Quadra", "Placar", "Status"
    ]
    ws_master.append(headers_master)
    ws_master.row_dimensions[2].height = 26
    
    for col_idx, text in enumerate(headers_master, start=1):
        cell = ws_master.cell(row=2, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill_blue
        cell.alignment = align_center
        cell.border = border_all

    m_row_idx = 3

    def escrever_linha_jogo(ws, r_idx, num_jogo, mod_nome, fase_nome, grupo_nome, rodada_val, time_a_str, time_b_str, data_str, hora_str, local_str, placar_str, status_str):
        row_values = [
            num_jogo, mod_nome, fase_nome, grupo_nome, rodada_val,
            time_a_str, time_b_str,
            data_str, hora_str, local_str, placar_str, status_str
        ]
        ws.append(row_values)
        ws.row_dimensions[r_idx].height = 22
        
        z_fill = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for c_i in range(1, 13):
            c = ws.cell(row=r_idx, column=c_i)
            c.font = font_regular
            c.border = border_all
            if z_fill.fill_type:
                c.fill = z_fill
            
            if c_i in [1, 3, 5, 8, 9, 11, 12]:
                c.alignment = align_center
            else:
                c.alignment = align_left
                
            if c_i in [8, 9, 10] and not row_values[c_i - 1]:
                c.font = font_small

    for mod, ch, partidas_mod, jogos_diretos in modalidades_com_partidas:
        for p in partidas_mod:
            jogo_obj = p.jogo
            
            num_jogo = f"J-{jogo_obj.id}" if jogo_obj else f"P-{p.id}"
            mod_nome = f"{mod.nome} ({mod.get_genero_display()})"
            fase_nome = p.get_fase_display()
            grupo_nome = p.grupo.nome if p.grupo else p.get_fase_display()
            rodada_val = p.rodada
            
            time_a_str = formatar_time_display(p.time_a or (jogo_obj.time_a if jogo_obj else None), p, 'a')
            time_b_str = formatar_time_display(p.time_b or (jogo_obj.time_b if jogo_obj else None), p, 'b')
            
            data_str = p.data_jogo_exibicao
            hora_str = p.horario_jogo_exibicao
            local_str = (jogo_obj.local if jogo_obj and jogo_obj.local else "")
            
            if p.placar_a is not None and p.placar_b is not None:
                placar_str = f"{p.placar_a} x {p.placar_b}"
            elif jogo_obj and jogo_obj.placar_time_a is not None and jogo_obj.placar_time_b is not None:
                placar_str = f"{jogo_obj.placar_time_a} x {jogo_obj.placar_time_b}"
            else:
                placar_str = "-"
                
            if p.finalizada or (jogo_obj and jogo_obj.finalizado):
                status_str = "Finalizado"
            elif data_str or hora_str:
                status_str = "Agendado"
            else:
                status_str = "Pendente"
                
            escrever_linha_jogo(
                ws_master, m_row_idx, num_jogo, mod_nome, fase_nome, grupo_nome,
                rodada_val, time_a_str, time_b_str, data_str, hora_str,
                local_str, placar_str, status_str
            )
            m_row_idx += 1
            
        for j in jogos_diretos:
            num_jogo = f"J-{j.id}"
            mod_nome = f"{mod.nome} ({mod.get_genero_display()})"
            fase_nome = "Partida Direta"
            grupo_nome = "Geral"
            rodada_val = 1
            
            time_a_str = j.time_a_display
            time_b_str = j.time_b_display
            
            data_str = j.data_jogo.strftime('%d/%m/%Y') if j.data_jogo else ""
            hora_str = j.horario_jogo.strftime('%H:%M') if j.horario_jogo else ""
            local_str = j.local or ""
            
            if j.placar_time_a is not None and j.placar_time_b is not None:
                placar_str = f"{j.placar_time_a} x {j.placar_time_b}"
            else:
                placar_str = "-"
                
            status_str = "Finalizado" if j.finalizado else ("Agendado" if data_str else "Pendente")
            
            escrever_linha_jogo(
                ws_master, m_row_idx, num_jogo, mod_nome, fase_nome, grupo_nome,
                rodada_val, time_a_str, time_b_str, data_str, hora_str,
                local_str, placar_str, status_str
            )
            m_row_idx += 1

    # --- 3. CONSTRUÇÃO DAS ABAS INDIVIDUAIS POR MODALIDADE ---
    for mod, ch, partidas_mod, jogos_diretos in modalidades_com_partidas:
        raw_name = f"{mod.nome} ({mod.get_genero_display()[:3]})"
        safe_sheet_title = "".join(c for c in raw_name if c not in r"\/*?:[]")[:30]
        
        ws_mod = wb.create_sheet(title=safe_sheet_title)
        ws_mod.views.sheetView[0].showGridLines = True
        
        ws_mod.merge_cells("A1:K1")
        mod_t = ws_mod["A1"]
        mod_t.value = f"JOGOS: {mod.nome.upper()} ({mod.get_genero_display().upper()})"
        mod_t.font = font_title
        mod_t.fill = header_fill_dark
        mod_t.alignment = align_center
        ws_mod.row_dimensions[1].height = 36
        
        headers_mod = [
            "# Jogo", "Fase", "Grupo / Chave", "Rodada",
            "Time A (Delegação)", "Time B (Delegação)",
            "Data do Jogo", "Horário", "Local / Quadra", "Placar", "Status"
        ]
        ws_mod.append(headers_mod)
        ws_mod.row_dimensions[2].height = 26
        
        for col_idx, text in enumerate(headers_mod, start=1):
            cell = ws_mod.cell(row=2, column=col_idx)
            cell.font = font_header
            cell.fill = header_fill_blue
            cell.alignment = align_center
            cell.border = border_all
            
        mod_row_idx = 3
        
        for p in partidas_mod:
            jogo_obj = p.jogo
            num_jogo = f"J-{jogo_obj.id}" if jogo_obj else f"P-{p.id}"
            fase_nome = p.get_fase_display()
            grupo_nome = p.grupo.nome if p.grupo else p.get_fase_display()
            rodada_val = p.rodada
            
            time_a_str = formatar_time_display(p.time_a or (jogo_obj.time_a if jogo_obj else None), p, 'a')
            time_b_str = formatar_time_display(p.time_b or (jogo_obj.time_b if jogo_obj else None), p, 'b')
            
            data_str = p.data_jogo_exibicao
            hora_str = p.horario_jogo_exibicao
            local_str = (jogo_obj.local if jogo_obj and jogo_obj.local else "")
            
            if p.placar_a is not None and p.placar_b is not None:
                placar_str = f"{p.placar_a} x {p.placar_b}"
            elif jogo_obj and jogo_obj.placar_time_a is not None and jogo_obj.placar_time_b is not None:
                placar_str = f"{jogo_obj.placar_time_a} x {jogo_obj.placar_time_b}"
            else:
                placar_str = "-"
                
            status_str = "Finalizado" if (p.finalizada or (jogo_obj and jogo_obj.finalizado)) else ("Agendado" if data_str else "Pendente")
            
            row_values = [
                num_jogo, fase_nome, grupo_nome, rodada_val,
                time_a_str, time_b_str, data_str, hora_str,
                local_str, placar_str, status_str
            ]
            ws_mod.append(row_values)
            ws_mod.row_dimensions[mod_row_idx].height = 22
            
            z_fill = zebra_fill if mod_row_idx % 2 == 0 else PatternFill(fill_type=None)
            for c_i in range(1, 12):
                c = ws_mod.cell(row=mod_row_idx, column=c_i)
                c.font = font_regular
                c.border = border_all
                if z_fill.fill_type:
                    c.fill = z_fill
                if c_i in [1, 2, 4, 7, 8, 10, 11]:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            mod_row_idx += 1

        for j in jogos_diretos:
            num_jogo = f"J-{j.id}"
            fase_nome = "Partida Direta"
            grupo_nome = "Geral"
            rodada_val = 1
            
            time_a_str = j.time_a_display
            time_b_str = j.time_b_display
            
            data_str = j.data_jogo.strftime('%d/%m/%Y') if j.data_jogo else ""
            hora_str = j.horario_jogo.strftime('%H:%M') if j.horario_jogo else ""
            local_str = j.local or ""
            
            placar_str = f"{j.placar_time_a} x {j.placar_time_b}" if (j.placar_time_a is not None and j.placar_time_b is not None) else "-"
            status_str = "Finalizado" if j.finalizado else ("Agendado" if data_str else "Pendente")
            
            row_values = [
                num_jogo, fase_nome, grupo_nome, rodada_val,
                time_a_str, time_b_str, data_str, hora_str,
                local_str, placar_str, status_str
            ]
            ws_mod.append(row_values)
            ws_mod.row_dimensions[mod_row_idx].height = 22
            
            z_fill = zebra_fill if mod_row_idx % 2 == 0 else PatternFill(fill_type=None)
            for c_i in range(1, 12):
                c = ws_mod.cell(row=mod_row_idx, column=c_i)
                c.font = font_regular
                c.border = border_all
                if z_fill.fill_type:
                    c.fill = z_fill
                if c_i in [1, 2, 4, 7, 8, 10, 11]:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            mod_row_idx += 1

    # --- 4. AUTO-AJUSTE DE LARGURA DE COLUNAS ---
    for ws in wb.worksheets:
        ws.freeze_panes = "A3"
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2]:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
